#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
從SME資料檔案得出. 中小企業在

1. 每一個縣市分布情形 (除批發零售業是小類, 其他用產業中類分別)
source file: sme_all_industry.csv

2. 根據業務區中心的分布情形 (除批發零售業是小類, 其他用產業中類分別)
source file: sme_all_industry.csv and 新興業務區中心.xlsx

3. 核貸案件在每一個縣市分布情形 (除批發零售業是小類, 其他用產業中類分別)
source file: uniaprcust.csv

4. 核准案件在每一個業務區中心的分布情形 (除批發零售業是小類, 其他用產業中類分別)
source file: uniaprcust.csv and 新興業務區中心.xlsx
    
Created on Mon May 15 15:52:40 2017

@author: fuming.tsai
"""
import datetime
from utils import getTmpFileNamePostfix
from openpyxl import Workbook

def printDictionary(result):
# just a test.
# print the result dictionary
    print('==================================================')
    print(result)

def ingestData(key, category, resultDictionary):    
    if key in resultDictionary:
        value=resultDictionary[key]
        if category in value:
            value[category]=value[category]+1
        else:
            value[category]=1
    else:
        value={}
        value[category]=1
        resultDictionary[key]=value
def outputStatisticsNumberToExcel(dataDictionary,areaSet, categorySet, tabTitle,fileName):
    rowFieldTitle=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    wb=Workbook()
    ws=wb.active
    ws.title=tabTitle
    #ws=wb.create_sheet(tabTitle)
    colIdx=2
    
    # add area as title
    for area in areaSet:
        ws.cell(row=1, column=colIdx, value=area)
        colIdx+=1
    ws.cell(row=1, column=colIdx, value='總計')
    
    rowIdx=2
    
    for category in categorySet:
        colIdx=2
        ws.cell(row=rowIdx, column=1, value=category)
        for area in areaSet:
            if area in dataDictionary and category in dataDictionary[area]:
                ws.cell(row=rowIdx, column=colIdx, value=dataDictionary[area][category])
            else:
                ws.cell(row=rowIdx, column=colIdx, value=0)
            colIdx+=1
        ws.cell(row=rowIdx, column=colIdx, value='=sum(B'+str(rowIdx)+':'+rowFieldTitle[colIdx-2]+str(rowIdx)+')')
       
        rowIdx+=1
    colIdx=2

    wb.save(fileName)
    print('Export the data to the file: '+fileName )


# I: setting
path='D:/fuming.Tsai/Documents/Tools/PortableGit/projects/GitDocs/06-專案文件/05-SME授信客戶資金需求/'
smeFile='sme_all_industry.csv'
#smeFile='test20170516.csv'
smeCategoryFile='新興業務區中心.csv'
reviewResult='uniaprcust.csv'
categoryThree=['批發及零售業']
outputFileCountyDistribution='county_stat_'+getTmpFileNamePostfix()+'.xlsx'
outputFileCallCenterDistribution='call_center_stat_'+getTmpFileNamePostfix()+'.xlsx'

# II: initialize
tmpDictionary={}
countyDictionary={}
callCenterDictionary={}
statisticsCountres={}
statisticsCallCenters={}
statisticsReviewCountres={}
statisticsReviewCallCenters={}
tmpFile=getTmpFileNamePostfix()+'.csv'
categorylist=set()


# get county data and district data
countyData=open(path+'/'+smeCategoryFile, 'r', encoding='UTF-8')
for record in countyData:
    data=record.split(',')
    countyDictionary[data[0].replace('\ufeff','')]=data[2].replace('\ufeff','')
    callCenterDictionary[data[0].replace('\ufeff','')]=data[1].replace('\ufeff','')
countyData.close()

# generate statistics report
smeSourceData=open(path+'/'+smeFile, 'r', encoding = 'UTF-8')
#smeTmpData=open(path+'/'+tmpFile, 'w',  encoding = 'UTF-8')
totalSMECount=0
for record in smeSourceData:
    data=record.split(',')
    isSME=data[13]
    totalSMECount+=1
    if 'Y' in isSME or 'y' in isSME: # for SME   
        #category=data[6]
        categoryOne=data[5]
        categorylist.add(categoryOne)
        #if categoryOne in categoryThree:
        #    category=data[7]      
        dataCounty=data[9]
        
        #  Part I: get category in each county
        statisticsCounty=countyDictionary[dataCounty]
        ingestData(statisticsCounty, categoryOne,statisticsCountres )
                              
        # Part II: get category in each call center  
        callCenter=callCenterDictionary[dataCounty]
        ingestData(callCenter, categoryOne,statisticsCallCenters )
           
#smeTmpData.close()
smeSourceData.close()

smeReviewData=open(path+'/'+reviewResult, 'r', encoding='UTF-8')
idx=0
for record in smeReviewData:
    if idx > 0:
        data=record.split(',')
        category=data[6]
        categoryOne=data[5]
        if categoryOne in categoryThree:
            category=data[7]      
        dataCounty=data[2]
        
        #  Part I: get category in each county
        statisticsCounty=countyDictionary[dataCounty]
        ingestData(statisticsCounty, categoryOne,statisticsReviewCountres )
                              
        # Part II: get category in each call center  
        callCenter=callCenterDictionary[dataCounty]
        ingestData(callCenter, categoryOne,statisticsReviewCallCenters )
    idx+=1
smeReviewData.close()

# (dataDictionary,areaSet, categorySet, tabTitle,fileName):
countyList=['基隆市','台北市','新北市','桃園市','新竹縣','新竹市','苗栗縣','台中市',
          '彰化縣','雲林縣','嘉義縣','嘉義市','台南市','高雄市','屏東縣','南投縣',
          '宜蘭縣','花蓮縣','台東縣','澎湖縣','金門縣','連江縣']

outputStatisticsNumberToExcel(dataDictionary=statisticsReviewCallCenters,
                              areaSet=statisticsReviewCallCenters.keys(), 
                             categorySet=categorylist,
                             tabTitle='放款名單家數分布',
                             fileName=path+'/放款名單家數分布_區域中心'+'.xlsx' )
outputStatisticsNumberToExcel(dataDictionary=statisticsReviewCountres,
                              areaSet=countyList, 
                             categorySet=categorylist,
                             tabTitle='放款名單家數分布',
                             fileName=path+'/放款名單家數分布_縣市'+'.xlsx' )

outputStatisticsNumberToExcel(dataDictionary=statisticsCountres,
                              areaSet=countyList, 
                             categorySet=categorylist,
                             tabTitle='全台中小企業分布',
                             fileName=path+'/全台中小企業分布_縣市'+'.xlsx' )
outputStatisticsNumberToExcel(dataDictionary=statisticsCallCenters,
                              areaSet=statisticsCallCenters.keys(), 
                             categorySet=categorylist,
                             tabTitle='全台中小企業分布',
                             fileName=path+'/全台中小企業分布_區域中心'+'.xlsx' )
#
#outputStatisticsNumberToExcel(statisticsCountres,('產業別','核貸家數','中小企業總家數'),'',path+'/'+outputFileCountyDistribution)
#printDictionary(statisticsCountres)
#printDictionary(statisticsCallCenters)
#printDictionary(countyDictionary)
#printDictionary(callCenterDictionary)
#printDictionary(statisticsReviewCountres)
#printDictionary(statisticsReviewCallCenters)
#print('total records: '+str(totalSMECount))


